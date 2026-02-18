"""관리대장 Excel 파일에 제출물 처리 결과를 기록하는 어댑터."""

from __future__ import annotations

import time
from pathlib import Path

from app.core.entities import LedgerRecord


# 지연 import로 의존성 확인 시점을 실제 사용 시점으로 늦춘다.
def _load_openpyxl_module():
    """openpyxl 모듈을 안전하게 로드한다."""
    try:
        import openpyxl
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required: pip install openpyxl") from exc
    return openpyxl


class OpenpyxlLedgerRepository:
    """제출물 관리대장 시트에 행을 추가한다."""

    def __init__(
        self,
        workbook_path: Path,
        sheet_name: str,
        business_columns: list[str] | tuple[str, ...],
        system_columns: list[str] | tuple[str, ...],
    ):
        """워크북 경로, 시트명, 기록할 컬럼 구성을 초기화한다."""
        self._workbook_path = workbook_path
        self._sheet_name = sheet_name
        self._business_columns = [c.strip() for c in business_columns if str(c).strip()]
        self._system_columns = [c.strip() for c in system_columns if str(c).strip()]

    # 헤더 자동 보정 후 제출물 1건을 시트에 추가한다.
    def append_submission(self, record: LedgerRecord) -> None:
        """제출물 처리 결과 1행을 워크북에 기록한다."""
        openpyxl = _load_openpyxl_module()

        retries = 3
        delay = 0.5
        last_error: Exception | None = None

        for _ in range(retries):
            wb = None
            try:
                if self._workbook_path.exists():
                    wb = openpyxl.load_workbook(self._workbook_path)
                else:
                    self._workbook_path.parent.mkdir(parents=True, exist_ok=True)
                    wb = openpyxl.Workbook()

                ws = self._get_or_create_sheet(wb)
                required_headers = self._required_headers()
                header_map = self._ensure_headers(ws, required_headers)
                payload = self._build_payload(record)

                row_idx = ws.max_row + 1
                for header in required_headers:
                    ws.cell(row=row_idx, column=header_map[header], value=payload.get(header))

                wb.save(self._workbook_path)
                wb.close()
                return
            except Exception as exc:  # pragma: no cover
                # 파일 잠금 등 일시 오류를 고려해 지수 백오프로 재시도한다.
                last_error = exc
                time.sleep(delay)
                delay *= 2
            finally:
                if wb is not None:
                    try:
                        wb.close()
                    except Exception:
                        pass

        assert last_error is not None
        raise last_error

    # 대상 시트가 없으면 생성하고, 비어있는 기본 시트는 이름을 바꿔 재사용한다.
    def _get_or_create_sheet(self, workbook):
        """대상 시트를 조회하고 없으면 생성한다."""
        if self._sheet_name in workbook.sheetnames:
            return workbook[self._sheet_name]

        if len(workbook.sheetnames) == 1 and workbook.active.max_row <= 1 and workbook.active.max_column <= 1:
            ws = workbook.active
            ws.title = self._sheet_name
            return ws

        return workbook.create_sheet(self._sheet_name)

    # business/system 컬럼을 중복 없이 합쳐 최종 필요 헤더를 만든다.
    def _required_headers(self) -> list[str]:
        """기록에 필요한 최종 헤더 목록을 계산한다."""
        headers: list[str] = []
        for name in self._business_columns + self._system_columns:
            if name not in headers:
                headers.append(name)
        if not headers:
            raise ValueError("No ledger columns configured for submission sheet")
        return headers

    # 시트의 1행 헤더를 확인하고 누락 헤더를 자동으로 추가한다.
    def _ensure_headers(self, ws, required_headers: list[str]) -> dict[str, int]:
        """시트 헤더를 보정하고 header->column index 매핑을 반환한다."""
        raw_headers = [ws.cell(row=1, column=idx).value for idx in range(1, ws.max_column + 1)]
        existing = [str(value).strip() for value in raw_headers if value is not None and str(value).strip()]

        if not existing:
            for idx, header in enumerate(required_headers, start=1):
                ws.cell(row=1, column=idx, value=header)
        else:
            existing_set = set(existing)
            next_col = max(ws.max_column, len(existing)) + 1
            for header in required_headers:
                if header not in existing_set:
                    ws.cell(row=1, column=next_col, value=header)
                    next_col += 1
                    existing_set.add(header)

        final_headers = [
            str(ws.cell(row=1, column=idx).value).strip()
            for idx in range(1, ws.max_column + 1)
            if ws.cell(row=1, column=idx).value is not None and str(ws.cell(row=1, column=idx).value).strip()
        ]
        header_map = {name: idx + 1 for idx, name in enumerate(final_headers)}

        missing = [header for header in required_headers if header not in header_map]
        if missing:
            raise ValueError(f"Failed to ensure required headers: {missing}")

        return header_map

    # 레코드 객체를 컬럼명 기반 payload dict로 변환한다.
    def _build_payload(self, record: LedgerRecord) -> dict[str, object]:
        """LedgerRecord를 시트 입력용 payload로 변환한다."""
        payload: dict[str, object] = {}

        business_values = {
            "제출물명": record.original_filename,
            "제출물 형태": Path(record.original_filename).suffix.lstrip(".").lower() or None,
            "접수일": record.processed_at.split("T")[0],
        }

        system_values = {
            "phase": record.phase,
            "artifact_abbr": record.artifact_abbr,
            "classification_method": record.classification_method,
            "ai_confidence": record.ai_confidence,
            "result_path": record.result_path,
            "status": record.status,
            "error_message": record.error_message,
            "processed_at": record.processed_at,
        }

        for column in self._business_columns:
            payload[column] = business_values.get(column)

        for column in self._system_columns:
            payload[column] = system_values.get(column)

        return payload
