"""프로젝트 문서번호 기반 분류를 담당하는 독립 어댑터 모듈.

다른 프로젝트에서도 재사용할 수 있도록 core와 분리해 유지한다.
"""

from __future__ import annotations

import re
from pathlib import Path

from app.core.entities import ClassificationResult, Knowledge


# 지연 import로 의존성 체크를 실제 사용 시점으로 늦춘다.
def _load_openpyxl_module():
    """openpyxl 모듈을 안전하게 로드한다."""
    try:
        import openpyxl
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required: pip install openpyxl") from exc
    return openpyxl


# 문서번호/문서구분 문자열을 비교 가능한 형태로 정규화한다.
def _norm(value: str) -> str:
    """영숫자만 남긴 대문자 문자열로 정규화한다."""
    return re.sub(r"[^A-Za-z0-9]", "", value).upper()


class ProjectDocumentClassifier:
    """문서번호 매핑표(엑셀)를 이용해 artifact/phase를 결정한다."""

    def __init__(self, workbook_path: Path, sheet_name: str, column_title_doc_no: str, column_title_abbr: str):
        """문서목록 엑셀의 위치/시트/컬럼 정보를 초기화한다."""
        self._workbook_path = workbook_path
        self._sheet_name = sheet_name
        self._column_title_doc_no = column_title_doc_no
        self._column_title_abbr = column_title_abbr
        self._rows: list[tuple[str, str, str]] = []
        self._load_rows()

    # 엑셀에서 문서번호/문서구분 매핑 데이터를 메모리에 적재한다.
    def _load_rows(self) -> None:
        """분류에 필요한 문서번호 매핑 행을 로드한다."""
        if not self._workbook_path.exists():
            return

        openpyxl = _load_openpyxl_module()
        wb = openpyxl.load_workbook(self._workbook_path, data_only=True)
        try:
            if self._sheet_name not in wb.sheetnames:
                return
            ws = wb[self._sheet_name]

            headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
            header_map = {name: idx + 1 for idx, name in enumerate(headers) if name}

            if self._column_title_doc_no not in header_map or self._column_title_abbr not in header_map:
                return

            doc_col = header_map[self._column_title_doc_no]
            abbr_col = header_map[self._column_title_abbr]

            loaded: list[tuple[str, str, str]] = []
            for row_idx in range(2, ws.max_row + 1):
                doc_raw = ws.cell(row=row_idx, column=doc_col).value
                abbr_raw = ws.cell(row=row_idx, column=abbr_col).value
                if doc_raw is None or abbr_raw is None:
                    continue

                doc_no = str(doc_raw).strip()
                abbr_text = str(abbr_raw).strip()
                if not doc_no or not abbr_text:
                    continue

                loaded.append((_norm(doc_no), doc_no, abbr_text))

            # 짧은 번호의 오매칭을 피하려고 긴 문서번호부터 우선 매칭한다.
            loaded.sort(key=lambda item: len(item[0]), reverse=True)
            self._rows = loaded
        finally:
            wb.close()

    # 파일명에 포함된 문서번호를 찾아 artifact/phase를 결정한다.
    def classify(self, filename: str, knowledge: Knowledge) -> ClassificationResult | None:
        """파일명에서 문서번호를 찾아 분류 결과를 반환한다."""
        if not self._rows:
            return None

        file_norm = _norm(filename)
        if not file_norm:
            return None

        best_match: tuple[str, str, str] | None = None
        for normalized_doc_no, raw_doc_no, abbr_text in self._rows:
            if normalized_doc_no and normalized_doc_no in file_norm:
                best_match = (normalized_doc_no, raw_doc_no, abbr_text)
                break

        if best_match is None:
            return None

        _, raw_doc_no, abbr_text = best_match
        resolved = self._resolve_artifact_from_abbr_text(abbr_text, knowledge)
        if resolved is None:
            return None

        artifact_abbr, phase = resolved
        return ClassificationResult(
            phase=phase,
            artifact_abbr=artifact_abbr,
            method="Rule-ABBR",
            confidence=1.0,
            reason=f"project_doc_info matched doc_no={raw_doc_no}",
        )

    @staticmethod
    # 문서구분 값은 postfix가 붙을 수 있으므로 유연 매칭한다.
    def _resolve_artifact_from_abbr_text(abbr_text: str, knowledge: Knowledge) -> tuple[str, str] | None:
        """문서구분 텍스트를 artifact 약어로 해석한다."""
        text_norm = _norm(abbr_text)
        if not text_norm:
            return None

        # 1) 정규화된 정확 일치
        for artifact in knowledge.artifacts:
            if _norm(artifact.abbr) == text_norm:
                return artifact.abbr, artifact.phase

        # 2) prefix/contains 유연 매칭 (postfix 허용)
        candidates: list[tuple[int, str, str]] = []
        for artifact in knowledge.artifacts:
            abbr_norm = _norm(artifact.abbr)
            if not abbr_norm:
                continue
            if text_norm.startswith(abbr_norm) or abbr_norm in text_norm:
                candidates.append((len(abbr_norm), artifact.abbr, artifact.phase))

        if not candidates:
            return None

        # 가장 긴 약어를 선택해 과도한 부분매칭을 줄인다.
        candidates.sort(reverse=True, key=lambda x: x[0])
        _, abbr, phase = candidates[0]
        return abbr, phase
