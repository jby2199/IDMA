from pathlib import Path

from app.adapters.excel_repo import OpenpyxlLedgerRepository
from app.core.entities import LedgerRecord


def _record() -> LedgerRecord:
    return LedgerRecord(
        original_filename="doc1.pdf",
        phase="Design",
        artifact_abbr="SDD",
        classification_method="Rule-ABBR",
        ai_confidence=None,
        result_path=r"C:\out\doc1.pdf",
        status="Filed",
        error_message="",
        processed_at="2026-02-15T10:00:00",
    )


def test_excel_repo_creates_headers_when_missing(tmp_path: Path) -> None:
    workbook = tmp_path / "ledger.xlsx"
    repo = OpenpyxlLedgerRepository(
        workbook_path=workbook,
        sheet_name="제출물 관리대장",
        business_columns=["제출물명", "제출물 형태", "접수일"],
        system_columns=["phase", "artifact_abbr", "status", "processed_at"],
    )

    repo.append_submission(_record())

    import openpyxl

    wb = openpyxl.load_workbook(workbook)
    ws = wb["제출물 관리대장"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    assert "제출물명" in headers
    assert "phase" in headers
    assert ws.max_row == 2

    header_map = {name: idx + 1 for idx, name in enumerate(headers)}
    assert ws.cell(2, header_map["제출물명"]).value == "doc1.pdf"
    assert ws.cell(2, header_map["phase"]).value == "Design"


def test_excel_repo_adds_missing_headers(tmp_path: Path) -> None:
    import openpyxl

    workbook = tmp_path / "ledger.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "제출물 관리대장"
    ws.cell(1, 1, "제출물명")
    wb.save(workbook)
    wb.close()

    repo = OpenpyxlLedgerRepository(
        workbook_path=workbook,
        sheet_name="제출물 관리대장",
        business_columns=["제출물명", "접수일"],
        system_columns=["phase", "status"],
    )

    repo.append_submission(_record())

    wb2 = openpyxl.load_workbook(workbook)
    ws2 = wb2["제출물 관리대장"]
    headers = [ws2.cell(1, c).value for c in range(1, ws2.max_column + 1)]
    assert "접수일" in headers
    assert "phase" in headers
    assert "status" in headers
